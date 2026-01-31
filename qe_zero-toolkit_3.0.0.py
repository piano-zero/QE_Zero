# QE Zero - Toolkit 
# Gestione Avanzata dei Quadri Economici Opere Pubbliche
# Versione 3.0.0
#
# Copyright (C) 2025 Rodolfo Sabelli
#
# Questo programma è software libero: puoi ridistribuirlo e/o modificarlo
# secondo i termini della GNU General Public License versione 3 o della
# European Union Public License versione 1.2 (a tua scelta).
#
# Questo programma è distribuito nella speranza che sia utile,
# ma SENZA ALCUNA GARANZIA; senza neppure la garanzia implicita di
# COMMERCIABILITÀ o IDONEITÀ PER UN PARTICOLARE SCOPO.
#
# Vedi LICENSE.txt per il testo completo delle licenze.
#
# Contatti: rodolfo.sabelli@gmail.com
# Repository: [URL GITHUB/GITLAB]

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import subprocess
import platform

# =============================================================================
# CLASSE TAB 1: ESPORTATORE SCHEDE CATALOGO
# =============================================================================
class TabExportCataloghi(ttk.Frame):
    def __init__(self, parent, db_path, app_root):
        super().__init__(parent)
        self.db_path = db_path
        self.app_root = app_root
        self.conn = None
        self.map_normative = {}
        self.normativa_selezionata_id = None
        self.setup_ui()
    
    def setup_ui(self):
        f_bot = ttk.Frame(self, padding=10)
        f_bot.pack(side='bottom', fill='x')
        
        ttk.Button(f_bot, text="Seleziona Tutto", command=self.seleziona_tutto).pack(side='left', padx=5)
        ttk.Button(f_bot, text="Esporta (Excel)", command=self.esporta_excel).pack(side='right', padx=5, fill='x', expand=True)

        paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashwidth=5, bg="#d9d9d9")
        paned.pack(side='top', fill='both', expand=True, padx=10, pady=10)

        f_left = ttk.LabelFrame(paned, text="1. Seleziona Catalogo", padding=5)
        paned.add(f_left, width=300)
        self.list_norm = tk.Listbox(f_left, selectmode=tk.SINGLE, font=("Segoe UI", 10))
        self.list_norm.pack(side='left', fill='both', expand=True)
        sb_norm = ttk.Scrollbar(f_left, orient="vertical", command=self.list_norm.yview)
        sb_norm.pack(side='right', fill='y')
        self.list_norm.config(yscrollcommand=sb_norm.set)
        self.list_norm.bind('<<ListboxSelect>>', self.on_select_normativa)

        f_right = ttk.LabelFrame(paned, text="2. Seleziona Voci", padding=5)
        paned.add(f_right)
        lbl_info = ttk.Label(f_right, text="CTRL+Click o SHIFT+Click per selezioni multiple", font=("Segoe UI", 8, "italic"))
        lbl_info.pack(fill='x', pady=(0,5))
        self.tree_voci = ttk.Treeview(f_right, columns=("Cod", "Desc"), show='headings', selectmode='extended')
        self.tree_voci.heading("Cod", text="Codice"); self.tree_voci.column("Cod", width=80, anchor='w')
        self.tree_voci.heading("Desc", text="Descrizione"); self.tree_voci.column("Desc", width=400, anchor='w')
        sb_voci = ttk.Scrollbar(f_right, orient="vertical", command=self.tree_voci.yview)
        self.tree_voci.configure(yscrollcommand=sb_voci.set)
        self.tree_voci.pack(side='left', fill='both', expand=True); sb_voci.pack(side='right', fill='y')

    def connetti_e_carica(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            cursor = self.conn.cursor()
            cursor.execute("SELECT id, nome FROM normative ORDER BY id")
            self.list_norm.delete(0, tk.END); self.map_normative = {}
            for idx, row in enumerate(cursor.fetchall()):
                self.list_norm.insert(tk.END, row[1])
                self.map_normative[idx] = row[0]
        except Exception as e:
            messagebox.showerror("Errore DB", str(e))

    def on_select_normativa(self, event):
        selection = self.list_norm.curselection()
        if not selection: return
        self.normativa_selezionata_id = self.map_normative[selection[0]]
        self.tree_voci.delete(*self.tree_voci.get_children())
        cur = self.conn.cursor()
        cur.execute("SELECT codice, descrizione FROM catalogo_voci WHERE normativa_id=? ORDER BY codice", (self.normativa_selezionata_id,))
        for row in cur.fetchall(): self.tree_voci.insert("", "end", values=row)

    def seleziona_tutto(self):
        for item in self.tree_voci.get_children(): self.tree_voci.selection_add(item)

    def pulisci_nome_foglio(self, testo):
        return re.sub(r'[\\/*?:\[\]]', '_', testo)[:31]

    def esporta_excel(self):
        sel = self.tree_voci.selection()
        if not sel: messagebox.showwarning("","Nessuna voce selezionata"); return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not fn: return
        try:
            wb = Workbook(); first = True
            for item in sel:
                v = self.tree_voci.item(item)['values']; cod = str(v[0]); desc = str(v[1])
                sh_name = self.pulisci_nome_foglio(cod)
                ws = wb.active if first else wb.create_sheet(title=sh_name)
                if first: ws.title = sh_name; first = False
                ws['A1'] = cod; ws['A1'].font = Font(bold=True, size=12)
                ws['A2'] = desc; ws['A2'].alignment = Alignment(wrap_text=True)
                ws.column_dimensions['A'].width = 50
            wb.save(fn)
            messagebox.showinfo("OK", "Export completato"); self.app_root.apri_file(fn)
        except Exception as e: messagebox.showerror("Errore", str(e))


# =============================================================================
# CLASSE TAB 2: RIEPILOGO BASE D'ASTA
# =============================================================================
class TabBaseAsta(ttk.Frame):
    def __init__(self, parent, db_path, app_root):
        super().__init__(parent)
        self.db_path = db_path
        self.app_root = app_root
        self.conn = None
        self.map_progetti = {} 
        self.map_qe = {}       
        self.data_cache = None 
        self.setup_ui()

    def setup_ui(self):
        f_sel = ttk.LabelFrame(self, text="Seleziona Fonte Dati", padding=10)
        f_sel.pack(side='top', fill='x', padx=10, pady=10)
        
        ttk.Label(f_sel, text="Progetto:").pack(side='left', padx=5)
        self.cb_prog = ttk.Combobox(f_sel, state="readonly", width=40)
        self.cb_prog.pack(side='left', padx=5)
        self.cb_prog.bind("<<ComboboxSelected>>", self.on_select_progetto)
        ttk.Label(f_sel, text="QE:").pack(side='left', padx=(20, 5))
        self.cb_qe = ttk.Combobox(f_sel, state="readonly", width=30)
        self.cb_qe.pack(side='left', padx=5)
        self.cb_qe.bind("<<ComboboxSelected>>", self.on_select_qe)

        f_bot = ttk.Frame(self, padding=10)
        f_bot.pack(side='bottom', fill='x')
        self.btn_export = ttk.Button(f_bot, text="Esporta (Excel)", command=self.esporta_excel, state='disabled')
        self.btn_export.pack(fill='x')

        f_prev = ttk.LabelFrame(self, text="Anteprima Dati (Layout Contabile)", padding=10)
        f_prev.pack(side='top', fill='both', expand=True, padx=10, pady=5)
        
        self.txt_preview = tk.Text(f_prev, font=("Consolas", 10), state='disabled', padx=15, pady=15, bg="white", wrap="none")
        self.txt_preview.pack(side='left', fill='both', expand=True)
        
        self.tabs_dati = (90, "left", 650, "right")
        self.tabs_header = (650, "right")

        self.txt_preview.tag_configure("head_blue", background="#DAE8FC", foreground="black", font=("Consolas", 11, "bold"), tabs=self.tabs_header)
        self.txt_preview.tag_configure("head_green", foreground="#009900", font=("Consolas", 11, "bold"), tabs=self.tabs_header)
        self.txt_preview.tag_configure("head_red", foreground="#FF0000", font=("Consolas", 11, "bold"), tabs=self.tabs_header)
        self.txt_preview.tag_configure("row", foreground="black", font=("Consolas", 10), tabs=self.tabs_dati)
        self.txt_preview.tag_configure("info", foreground="gray", font=("Consolas", 9, "italic"))

        sb = ttk.Scrollbar(f_prev, command=self.txt_preview.yview)
        sb.pack(side='right', fill='y')
        self.txt_preview['yscrollcommand'] = sb.set

    def connetti_e_carica(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            cur = self.conn.cursor()
            cur.execute("SELECT id, titolo, cup FROM progetti ORDER BY id DESC")
            rows = cur.fetchall()
            self.cb_prog['values'] = [f"{r[1]} (CUP: {r[2]})" for r in rows]
            self.map_progetti = {i: r[0] for i, r in enumerate(rows)}
        except Exception as e:
            messagebox.showerror("Errore DB", str(e))

    def on_select_progetto(self, e):
        idx = self.cb_prog.current()
        if idx == -1: return
        pid = self.map_progetti[idx]
        cur = self.conn.cursor()
        cur.execute("SELECT id, nome_versione FROM quadri_economici WHERE progetto_id=? ORDER BY id DESC", (pid,))
        rows = cur.fetchall()
        self.cb_qe['values'] = [r[1] for r in rows]
        self.cb_qe.set("")
        self.map_qe = {i: r[0] for i, r in enumerate(rows)}
        self.reset_preview()

    def reset_preview(self):
        self.txt_preview.config(state='normal'); self.txt_preview.delete("1.0", tk.END); self.txt_preview.config(state='disabled')
        self.btn_export.config(state='disabled'); self.data_cache = None

    def fmt(self, val):
        try: v = float(val)
        except: v = 0.0
        return f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def on_select_qe(self, e):
        idx = self.cb_qe.current()
        if idx == -1: return
        self.calcola_riepilogo(self.map_qe[idx])

    def calcola_riepilogo(self, qid):
        cur = self.conn.cursor()
        try:
            cur.execute("""SELECT codice_completo, descrizione, valore_imponibile, is_percentuale, 
                           flag_base_asta, flag_soggetto_ribasso, flag_calcolo_montante
                           FROM voci WHERE qe_id=? ORDER BY codice_completo ASC""", (qid,))
            rows = cur.fetchall()
        except sqlite3.OperationalError:
            cur.execute("""SELECT codice_completo, descrizione, valore_imponibile, is_percentuale, 
                           flag_base_asta, flag_soggetto_ribasso FROM voci WHERE qe_id=? ORDER BY codice_completo ASC""", (qid,))
            rows = [list(r) + [0] for r in cur.fetchall()]

        montante = 0.0
        for r in rows:
            if r[3] == 0 and r[6] == 1: montante += r[2]

        lista_A = []; lista_B = []; tot_A = 0.0; tot_B = 0.0
        
        for r in rows:
            cod, desc, raw_val, is_perc, flg_base, flg_rib, _ = r
            if flg_base == 1: 
                imp = raw_val if is_perc == 0 else (montante * raw_val / 100)
                item = (cod, desc, imp)
                if flg_rib == 1: lista_A.append(item); tot_A += imp
                else: lista_B.append(item); tot_B += imp
        
        tot_gen = tot_A + tot_B
        self.data_cache = {"tot_gen": tot_gen, "tot_A": tot_A, "lista_A": lista_A, "tot_B": tot_B, "lista_B": lista_B}
        self.mostra_anteprima()

    def mostra_anteprima(self):
        self.txt_preview.config(state='normal')
        self.txt_preview.delete("1.0", tk.END)
        self.txt_preview.insert(tk.END, f"IMPORTO TOTALE A BASE D'ASTA\t{self.fmt(self.data_cache['tot_gen'])}\n", "head_blue")
        self.txt_preview.insert(tk.END, "\n")
        self.txt_preview.insert(tk.END, f"A) IMPORTO SOGGETTO A RIBASSO\t{self.fmt(self.data_cache['tot_A'])}\n", "head_green")
        if not self.data_cache['lista_A']:
            self.txt_preview.insert(tk.END, "   (Nessuna voce presente)\n", "info")
        else:
            for item in self.data_cache['lista_A']:
                desc = (item[1][:70] + '..') if len(item[1]) > 70 else item[1]
                riga = f"{item[0]}\t{desc}\t{self.fmt(item[2])}\n"
                self.txt_preview.insert(tk.END, riga, "row")
        self.txt_preview.insert(tk.END, "\n")
        self.txt_preview.insert(tk.END, f"B) SOMME NON SOGGETTE A RIBASSO\t{self.fmt(self.data_cache['tot_B'])}\n", "head_red")
        if not self.data_cache['lista_B']:
            self.txt_preview.insert(tk.END, "   (Nessuna voce presente)\n", "info")
        else:
            for item in self.data_cache['lista_B']:
                desc = (item[1][:70] + '..') if len(item[1]) > 70 else item[1]
                riga = f"{item[0]}\t{desc}\t{self.fmt(item[2])}\n"
                self.txt_preview.insert(tk.END, riga, "row")
        self.txt_preview.config(state='disabled')
        self.btn_export.config(state='normal')

    def esporta_excel(self):
        if not self.data_cache: return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Salva Riepilogo")
        if not fn: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Base d'Asta"
            f_head = Font(name="Arial", size=11, bold=True)
            f_std = Font(name="Arial", size=11)
            f_green = Font(name="Arial", size=11, bold=True, color="009900")
            f_red = Font(name="Arial", size=11, bold=True, color="FF0000")
            fill_blue = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
            b_bot = Border(bottom=Side(style='thin'))
            a_right = Alignment(horizontal="right"); a_left = Alignment(horizontal="left")
            
            ws.merge_cells('A1:B1')
            ws['A1'] = "IMPORTO TOTALE A BASE D'ASTA"; ws['C1'] = float(self.data_cache['tot_gen'])
            ws['A1'].font = f_head; ws['A1'].fill = fill_blue; ws['B1'].fill = fill_blue
            ws['C1'].font = f_head; ws['C1'].fill = fill_blue; ws['C1'].number_format = '#,##0.00 €'; ws['C1'].alignment = a_right
            
            row = 3
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = "A) IMPORTO SOGGETTO A RIBASSO"; ws[f'C{row}'] = float(self.data_cache['tot_A'])
            ws[f'A{row}'].font = f_green; ws[f'C{row}'].font = f_green; ws[f'C{row}'].number_format = '#,##0.00 €'; ws[f'C{row}'].alignment = a_right
            row += 1
            for i in self.data_cache['lista_A']:
                ws[f'A{row}'] = i[0]; ws[f'B{row}'] = i[1]; ws[f'C{row}'] = float(i[2])
                ws[f'C{row}'].number_format = '#,##0.00 €'; ws[f'C{row}'].alignment = a_right
                for c in ['A','B','C']: ws[f'{c}{row}'].font = f_std; ws[f'{c}{row}'].border = b_bot
                row += 1
            row += 2
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = "B) SOMME NON SOGGETTE A RIBASSO"; ws[f'C{row}'] = float(self.data_cache['tot_B'])
            ws[f'A{row}'].font = f_red; ws[f'C{row}'].font = f_red; ws[f'C{row}'].number_format = '#,##0.00 €'; ws[f'C{row}'].alignment = a_right
            row += 1
            for i in self.data_cache['lista_B']:
                ws[f'A{row}'] = i[0]; ws[f'B{row}'] = i[1]; ws[f'C{row}'] = float(i[2])
                ws[f'C{row}'].number_format = '#,##0.00 €'; ws[f'C{row}'].alignment = a_right
                for c in ['A','B','C']: ws[f'{c}{row}'].font = f_std; ws[f'{c}{row}'].border = b_bot
                row += 1
            ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 75; ws.column_dimensions['C'].width = 25
            wb.save(fn)
            messagebox.showinfo("Export", "File creato!"); self.app_root.apri_file(fn)
        except Exception as e: messagebox.showerror("Errore", str(e))

# =============================================================================
# CLASSE TAB 3: RIEPILOGO IVA (LAYOUT CORRETTO 2 COLONNE)
# =============================================================================
class TabRiepilogoIva(ttk.Frame):
    def __init__(self, parent, db_path, app_root):
        super().__init__(parent)
        self.db_path = db_path
        self.app_root = app_root
        self.conn = None
        self.map_progetti = {} 
        self.map_qe = {}        
        self.dati_iva = None 
        self.setup_ui()

    def setup_ui(self):
        f_sel = ttk.LabelFrame(self, text="Seleziona Fonte Dati", padding=10)
        f_sel.pack(side='top', fill='x', padx=10, pady=10)
        
        ttk.Label(f_sel, text="Progetto:").pack(side='left', padx=5)
        self.cb_prog = ttk.Combobox(f_sel, state="readonly", width=40)
        self.cb_prog.pack(side='left', padx=5)
        self.cb_prog.bind("<<ComboboxSelected>>", self.on_select_progetto)
        
        ttk.Label(f_sel, text="QE:").pack(side='left', padx=(20, 5))
        self.cb_qe = ttk.Combobox(f_sel, state="readonly", width=30)
        self.cb_qe.pack(side='left', padx=5)
        self.cb_qe.bind("<<ComboboxSelected>>", self.on_select_qe)

        f_bot = ttk.Frame(self, padding=10)
        f_bot.pack(side='bottom', fill='x')
        self.btn_export = ttk.Button(f_bot, text="Esporta (Excel)", command=self.esporta_excel, state='disabled')
        self.btn_export.pack(fill='x')

        f_prev = ttk.LabelFrame(self, text="Riepilogo Rendicontazione IVA", padding=10)
        f_prev.pack(side='top', fill='both', expand=True, padx=10, pady=5)
        
        self.txt_preview = tk.Text(f_prev, font=("Consolas", 10), state='disabled', padx=15, pady=15, bg="white", wrap="word")
        self.txt_preview.pack(side='left', fill='both', expand=True)
        
        self.tabs_header = (650, "right")
        self.txt_preview.tag_configure("head_blue", background="#DAE8FC", foreground="black", font=("Consolas", 11, "bold"), tabs=self.tabs_header)
        self.txt_preview.tag_configure("row_green", foreground="#009900", font=("Consolas", 10), tabs=self.tabs_header)
        self.txt_preview.tag_configure("row_red", foreground="#FF0000", font=("Consolas", 10), tabs=self.tabs_header)
        self.txt_preview.tag_configure("note_header", foreground="black", font=("Consolas", 10, "bold"))
        self.txt_preview.tag_configure("note_body", foreground="#333333", font=("Consolas", 9))
        self.txt_preview.tag_configure("info", foreground="gray", font=("Consolas", 9, "italic"))

        sb = ttk.Scrollbar(f_prev, command=self.txt_preview.yview)
        sb.pack(side='right', fill='y')
        self.txt_preview['yscrollcommand'] = sb.set

    def connetti_e_carica(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            cur = self.conn.cursor()
            cur.execute("SELECT id, titolo, cup FROM progetti ORDER BY id DESC")
            rows = cur.fetchall()
            self.cb_prog['values'] = [f"{r[1]} (CUP: {r[2]})" for r in rows]
            self.map_progetti = {i: r[0] for i, r in enumerate(rows)}
        except Exception as e:
            messagebox.showerror("Errore DB", str(e))

    def on_select_progetto(self, e):
        idx = self.cb_prog.current()
        if idx == -1: return
        pid = self.map_progetti[idx]
        cur = self.conn.cursor()
        cur.execute("SELECT id, nome_versione FROM quadri_economici WHERE progetto_id=? ORDER BY id DESC", (pid,))
        rows = cur.fetchall()
        self.cb_qe['values'] = [r[1] for r in rows]
        self.cb_qe.set("")
        self.map_qe = {i: r[0] for i, r in enumerate(rows)}
        self.reset_preview()

    def reset_preview(self):
        self.txt_preview.config(state='normal')
        self.txt_preview.delete("1.0", tk.END)
        self.txt_preview.config(state='disabled')
        self.btn_export.config(state='disabled')
        self.dati_iva = None

    def fmt(self, val):
        try: v = float(val)
        except: v = 0.0
        return f"€ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def on_select_qe(self, e):
        idx = self.cb_qe.current()
        if idx == -1: return
        self.calcola_iva(self.map_qe[idx])

    def calcola_iva(self, qid):
        cur = self.conn.cursor()
        query = """
            SELECT 
                valore_imponibile,      -- 0
                is_percentuale,         -- 1
                flag_calcolo_montante,  -- 2
                flag_base_asta,         -- 3
                perc_iva,               -- 4
                perc_oneri,             -- 5
                includi_oneri_in_iva,   -- 6
                codice_completo         -- 7
            FROM voci 
            WHERE qe_id=?
            ORDER BY codice_completo
        """
        try:
            cur.execute(query, (qid,))
            rows = cur.fetchall()
        except Exception as e:
            messagebox.showerror("Errore SQL", f"Errore lettura voci:\n{e}")
            return

        montante = sum(r[0] for r in rows if r[1] == 0 and r[2] == 1)

        iva_base_dict = {} 
        iva_oneri_dict = {}
        totale_generale_iva = 0.0

        for r in rows:
            raw_val = r[0]
            is_perc = r[1]
            codice = r[7] if r[7] else "?"
            try: aliquota_iva = float(r[4]) if r[4] is not None else 0.0
            except: aliquota_iva = 0.0
            try: perc_oneri = float(r[5]) if r[5] is not None else 0.0
            except: perc_oneri = 0.0
            flag_iva_su_oneri = (r[6] == 1) 

            valore_riga = (montante * raw_val / 100) if is_perc == 1 else raw_val

            if aliquota_iva > 0:
                iva_base_calc = valore_riga * (aliquota_iva / 100.0)
                if iva_base_calc != 0:
                    if aliquota_iva not in iva_base_dict:
                        iva_base_dict[aliquota_iva] = {'importo': 0.0, 'codici': []}
                    iva_base_dict[aliquota_iva]['importo'] += iva_base_calc
                    iva_base_dict[aliquota_iva]['codici'].append(str(codice))
                    totale_generale_iva += iva_base_calc

            if perc_oneri > 0 and flag_iva_su_oneri and aliquota_iva > 0:
                valore_onere = valore_riga * (perc_oneri / 100.0)
                iva_onere_calc = valore_onere * (aliquota_iva / 100.0)
                if iva_onere_calc != 0:
                    if aliquota_iva not in iva_oneri_dict:
                        iva_oneri_dict[aliquota_iva] = {'importo': 0.0, 'codici': []}
                    iva_oneri_dict[aliquota_iva]['importo'] += iva_onere_calc
                    iva_oneri_dict[aliquota_iva]['codici'].append(str(codice))
                    totale_generale_iva += iva_onere_calc

        self.dati_iva = {
            "base": iva_base_dict,
            "oneri": iva_oneri_dict,
            "totale": totale_generale_iva
        }
        self.mostra_risultati()

    def get_perc_label(self, val):
        return f"{int(val)}" if val.is_integer() else f"{val:.2f}"

    def mostra_risultati(self):
        self.txt_preview.config(state='normal')
        self.txt_preview.delete("1.0", tk.END)
        dati = self.dati_iva
        
        self.txt_preview.insert(tk.END, f"TOTALE IVA CALCOLATA\t{self.fmt(dati['totale'])}\n", "head_blue")
        self.txt_preview.insert(tk.END, "\n")
        
        note_map = {}
        note_counter = 1
        
        keys_base = sorted(dati["base"].keys())
        if keys_base:
            for k in keys_base:
                item = dati["base"][k]
                note_map[note_counter] = item['codici']
                desc = f"IVA al {self.get_perc_label(k)}% su Imponibile [{note_counter}]"
                self.txt_preview.insert(tk.END, f"{desc}\t{self.fmt(item['importo'])}\n", "row_green")
                note_counter += 1
        else:
             self.txt_preview.insert(tk.END, "(Nessuna IVA su imponibile)\n", "info")

        self.txt_preview.insert(tk.END, "\n")

        keys_oneri = sorted(dati["oneri"].keys())
        if keys_oneri:
            for k in keys_oneri:
                item = dati["oneri"][k]
                note_map[note_counter] = item['codici']
                desc = f"IVA al {self.get_perc_label(k)}% su Oneri e Imposte [{note_counter}]"
                self.txt_preview.insert(tk.END, f"{desc}\t{self.fmt(item['importo'])}\n", "row_red")
                note_counter += 1
        else:
             self.txt_preview.insert(tk.END, "(Nessuna IVA su oneri)\n", "info")

        self.txt_preview.insert(tk.END, "\n\n")
        self.txt_preview.insert(tk.END, "RIFERIMENTO VOCI:\n", "head_blue")
        self.txt_preview.insert(tk.END, "\n")
        
        for idx in sorted(note_map.keys()):
            codici_str = " - ".join(note_map[idx])
            self.txt_preview.insert(tk.END, f"[{idx}] ", "note_header")
            self.txt_preview.insert(tk.END, f"{codici_str}\n", "note_body")

        self.dati_iva['note_map'] = note_map
        self.txt_preview.config(state='disabled')
        self.txt_preview.update_idletasks()
        self.btn_export.config(state='normal')

    def esporta_excel(self):
        if not self.dati_iva: return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Salva Riepilogo IVA")
        if not fn: return
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Riepilogo IVA"
            
            f_head = Font(name="Arial", size=11, bold=True)
            f_green = Font(name="Arial", size=10, color="009900")
            f_red = Font(name="Arial", size=10, color="FF0000")
            f_note = Font(name="Arial", size=9, italic=True)
            
            fill_blue = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")
            b_bot = Border(bottom=Side(style='thin'))
            a_right = Alignment(horizontal="right", vertical="center")
            a_left = Alignment(horizontal="left", vertical="center")
            a_wrap = Alignment(wrap_text=True, vertical="top")

            ws['A1'] = "TOTALE IVA CALCOLATA"
            ws['B1'] = float(self.dati_iva['totale'])
            ws['A1'].font = f_head; ws['A1'].fill = fill_blue; ws['A1'].alignment = a_left
            ws['B1'].font = f_head; ws['B1'].fill = fill_blue; ws['B1'].alignment = a_right
            ws['B1'].number_format = '#,##0.00 €'

            ws.column_dimensions['A'].width = 70
            ws.column_dimensions['B'].width = 25

            row = 3
            note_counter = 1
            note_map_export = {}
            
            def scrivi_sezione(dict_dati, font_style, fmt_string):
                nonlocal row, note_counter
                for k in sorted(dict_dati.keys()):
                    item = dict_dati[k]
                    note_map_export[note_counter] = item['codici']
                    desc = fmt_string.format(perc=self.get_perc_label(k), note=note_counter)
                    
                    ws[f'A{row}'] = desc
                    ws[f'B{row}'] = float(item['importo'])
                    ws[f'A{row}'].font = font_style; ws[f'A{row}'].alignment = a_left; ws[f'A{row}'].border = b_bot
                    ws[f'B{row}'].font = font_style; ws[f'B{row}'].number_format = '#,##0.00 €'; ws[f'B{row}'].alignment = a_right; ws[f'B{row}'].border = b_bot
                    row += 1; note_counter += 1

            if self.dati_iva["base"]: scrivi_sezione(self.dati_iva["base"], f_green, "IVA al {perc}% su Imponibile [{note}]")
            if self.dati_iva["oneri"]: scrivi_sezione(self.dati_iva["oneri"], f_red, "IVA al {perc}% su Oneri e Imposte [{note}]")

            row += 2
            ws[f'A{row}'] = "RIFERIMENTO VOCI (NOTE):"; ws[f'A{row}'].font = Font(name="Arial", size=10, bold=True); row += 1
            
            for idx in sorted(note_map_export.keys()):
                codici_str = " - ".join(note_map_export[idx])
                ws.merge_cells(f'A{row}:B{row}')
                cell = ws[f'A{row}']
                cell.value = f"[{idx}] {codici_str}"
                cell.font = f_note; cell.alignment = a_wrap
                row += 1

            wb.save(fn)
            messagebox.showinfo("OK", "Export completato"); self.app_root.apri_file(fn)
        except Exception as e: messagebox.showerror("Errore Export", str(e))

# =============================================================================
# CLASSE TAB 4: CRONOPROGRAMMA A VERSIONI (MODIFICATA)
# =============================================================================
class TabCronoprogramma(ttk.Frame):
    def __init__(self, parent, db_path, app_root):
        super().__init__(parent)
        self.db_path = db_path
        self.app_root = app_root
        self.conn = None
        self.map_progetti = {} 
        self.map_qe = {}
        self.map_versioni = {} 
        
        self.voce_corrente_id = None 
        self.voce_corrente_iid = None 
        self.totale_voce_target = 0.0
        
        self.init_db_structure()
        self.setup_ui()

    def init_db_structure(self):
        try:
            conn = sqlite3.connect(self.db_path)
            conn.execute('''CREATE TABLE IF NOT EXISTS fpv_testata (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                qe_id INTEGER,
                descrizione TEXT,
                data_creazione TEXT,
                FOREIGN KEY (qe_id) REFERENCES quadri_economici (id) ON DELETE CASCADE
            )''')
            conn.execute('''CREATE TABLE IF NOT EXISTS fpv_dettaglio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                versione_id INTEGER,
                voce_id INTEGER,
                fornitore TEXT,
                anno_1 REAL,
                anno_2 REAL,
                anno_3 REAL,
                FOREIGN KEY (versione_id) REFERENCES fpv_testata (id) ON DELETE CASCADE,
                FOREIGN KEY (voce_id) REFERENCES voci (id) ON DELETE CASCADE
            )''')
            conn.commit(); conn.close()
        except Exception as e:
            messagebox.showerror("Errore DB", f"Impossibile inizializzare tabelle FPV:\n{e}")

    def setup_ui(self):
        f_sel = ttk.LabelFrame(self, text="1. Seleziona Progetto e QE", padding=10)
        f_sel.pack(side='top', fill='x', padx=10, pady=5)
        
        # Frame Progetto e QE 50/50
        f_p = ttk.Frame(f_sel)
        f_p.pack(side='left', fill='x', expand=True, padx=(0, 5))
        ttk.Label(f_p, text="Progetto:").pack(anchor='w')
        self.cb_prog = ttk.Combobox(f_p, state="readonly")
        self.cb_prog.pack(fill='x')
        self.cb_prog.bind("<<ComboboxSelected>>", self.on_select_progetto)

        f_q = ttk.Frame(f_sel)
        f_q.pack(side='left', fill='x', expand=True, padx=(5, 0))
        ttk.Label(f_q, text="QE:").pack(anchor='w')
        self.cb_qe = ttk.Combobox(f_q, state="readonly")
        self.cb_qe.pack(fill='x')
        self.cb_qe.bind("<<ComboboxSelected>>", self.on_select_qe)

        f_ver = ttk.LabelFrame(self, text="2. Seleziona Piano Finanziario (Versione)", padding=10)
        f_ver.pack(side='top', fill='x', padx=10, pady=5)
        ttk.Label(f_ver, text="Sorgente Dati:").pack(side='left', padx=5)
        self.cb_ver = ttk.Combobox(f_ver, state="readonly", width=70)
        self.cb_ver.pack(side='left', padx=5)
        self.cb_ver.bind("<<ComboboxSelected>>", self.carica_versione_selezionata)
        
        f_bot = ttk.Frame(self, padding=10)
        f_bot.pack(side='bottom', fill='x')
        ttk.Button(f_bot, text="💾 SALVA STATO ATTUALE COME NUOVA VERSIONE", command=self.salva_nuova_versione).pack(side='right', padx=5)
        ttk.Button(f_bot, text="Esporta (Excel)", command=self.esporta_excel).pack(side='right', padx=5)

        paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, sashwidth=5, bg="#d9d9d9")
        paned.pack(side='top', fill='both', expand=True, padx=10, pady=5)

        f_list = ttk.LabelFrame(paned, text="Dettaglio Voci", padding=5)
        paned.add(f_list, width=600)
        self.tr = ttk.Treeview(f_list, columns=("Cod", "Desc", "Fornitore", "Totale", "Stato", "A1", "A2", "A3"), show='headings', selectmode='browse')
        self.tr.heading("Cod", text="Cod"); self.tr.column("Cod", width=50)
        self.tr.heading("Desc", text="Descrizione"); self.tr.column("Desc", width=250)
        self.tr.heading("Fornitore", text="Beneficiario"); self.tr.column("Fornitore", width=150)
        self.tr.heading("Totale", text="Totale Lordo"); self.tr.column("Totale", width=90, anchor='e')
        self.tr.heading("Stato", text="Check"); self.tr.column("Stato", width=40, anchor='center')
        
        self.tr.column("A1", width=0, stretch=False)
        self.tr.column("A2", width=0, stretch=False)
        self.tr.column("A3", width=0, stretch=False)
        self.tr.tag_configure('ok', foreground='green'); self.tr.tag_configure('err', foreground='red')
        
        sb = ttk.Scrollbar(f_list, orient="vertical", command=self.tr.yview)
        self.tr.configure(yscrollcommand=sb.set)
        self.tr.pack(side='left', fill='both', expand=True); sb.pack(side='right', fill='y')
        self.tr.bind("<<TreeviewSelect>>", self.on_select_voce)

        f_edit = ttk.LabelFrame(paned, text="Modifica Voce Corrente", padding=15)
        paned.add(f_edit)
        self.lbl_info_voce = ttk.Label(f_edit, text="Seleziona una voce...", font=("Segoe UI", 9, "bold"), foreground="#555", wraplength=300)
        self.lbl_info_voce.pack(fill='x', pady=(0, 15))
        
        ttk.Label(f_edit, text="Soggetto / Fornitore:").pack(anchor='w')
        self.var_forn = tk.StringVar()
        ttk.Entry(f_edit, textvariable=self.var_forn).pack(fill='x', pady=(0, 10))

        f_grid = ttk.Frame(f_edit); f_grid.pack(fill='x', pady=5)
        self.vars_anni = [tk.StringVar(value="0,00") for _ in range(3)]
        labels = ["Anno 1:", "Anno 2:", "Anno 3 (Residuo):"]
        for i in range(3):
            ttk.Label(f_grid, text=labels[i]).grid(row=i, column=0, sticky='w', pady=5)
            e = ttk.Entry(f_grid, textvariable=self.vars_anni[i], justify='right')
            e.grid(row=i, column=1, sticky='ew', padx=10, pady=5)
            if i < 2: e.bind('<KeyRelease>', self.calcola_dinamica)
            else: e.config(state='readonly')

        f_grid.columnconfigure(1, weight=1)
        ttk.Separator(f_edit, orient='horizontal').pack(fill='x', pady=15)
        f_res = ttk.Frame(f_edit); f_res.pack(fill='x')
        self.lbl_diff = ttk.Label(f_res, text="...", font=("Segoe UI", 10, "bold")); self.lbl_diff.pack()
        
        # --- NUOVO BOTTONE: Sposta Anno 3 su Anno 1 ---
        ttk.Button(f_edit, text="⬆ Sposta Residuo (A3) su Anno 1", command=self.sposta_residuo_su_a1).pack(fill='x', pady=(15, 5))
        # ----------------------------------------------

        ttk.Button(f_edit, text="Salva Modifiche Riga", command=self.applica_modifiche_riga).pack(fill='x', pady=5)
        ttk.Label(f_edit, text="(Salva come Nuova Versione per confermare nel DB)", font=("Segoe UI", 8, "italic"), foreground="gray").pack()

    def fmt(self, val): 
        try: v = float(val)
        except: v = 0.0
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    def parse(self, val_str):
        if not val_str: return 0.0
        try: return float(str(val_str).replace("€", "").strip().replace(".", "").replace(",", "."))
        except: return 0.0

    def connetti_e_carica(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            cur = self.conn.cursor()
            cur.execute("SELECT id, titolo FROM progetti ORDER BY id DESC")
            rows = cur.fetchall()
            self.cb_prog['values'] = [r[1] for r in rows]
            self.map_progetti = {i: r[0] for i, r in enumerate(rows)}
        except Exception: pass

    def on_select_progetto(self, e):
        idx = self.cb_prog.current()
        if idx == -1: return
        pid = self.map_progetti[idx]
        cur = self.conn.cursor()
        cur.execute("SELECT id, nome_versione FROM quadri_economici WHERE progetto_id=? ORDER BY id DESC", (pid,))
        rows = cur.fetchall()
        self.cb_qe['values'] = [r[1] for r in rows]
        self.cb_qe.set("")
        self.map_qe = {i: r[0] for i, r in enumerate(rows)}
        self.tr.delete(*self.tr.get_children())
        self.cb_ver.set(""); self.cb_ver['values'] = []

    def on_select_qe(self, e):
        idx = self.cb_qe.current()
        if idx == -1: return
        self.refresh_versioni()

    def refresh_versioni(self):
        idx = self.cb_qe.current()
        if idx == -1: return
        qid = self.map_qe[idx]
        cur = self.conn.cursor()
        cur.execute("SELECT id, descrizione, data_creazione FROM fpv_testata WHERE qe_id=? ORDER BY id DESC", (qid,))
        rows = cur.fetchall()
        self.map_versioni = {}
        vals = ["✨ GENERA EX-NOVO (Usa dati attuali del QE)"]
        self.map_versioni[0] = None
        for i, r in enumerate(rows):
            vals.append(f"📁 Versione del {r[2]} - {r[1]}")
            self.map_versioni[i+1] = r[0]
        self.cb_ver['values'] = vals
        self.cb_ver.current(1 if rows else 0)
        self.carica_versione_selezionata(None)

    def carica_versione_selezionata(self, e):
        idx = self.cb_ver.current()
        qe_idx = self.cb_qe.current()
        if qe_idx == -1 or idx == -1: return
        qid = self.map_qe[qe_idx]
        ver_id = self.map_versioni.get(idx)
        self.carica_dati_base(qid, versione_id=ver_id)

    def carica_dati_base(self, qid, versione_id=None):
        self.tr.delete(*self.tr.get_children())
        self.pulisci_form()
        cur = self.conn.cursor()
        
        if versione_id:
            query = """SELECT v.id, v.codice_completo, v.descrizione, v.valore_imponibile, v.is_percentuale, v.flag_calcolo_montante,
                       v.perc_oneri, v.includi_oneri_in_iva, v.perc_iva, d.fornitore, d.anno_1, d.anno_2, d.anno_3
                       FROM voci v LEFT JOIN fpv_dettaglio d ON v.id = d.voce_id AND d.versione_id = ?
                       WHERE v.qe_id=? ORDER BY v.codice_completo"""
            cur.execute(query, (versione_id, qid))
        else:
            query = """SELECT id, codice_completo, descrizione, valore_imponibile, is_percentuale, flag_calcolo_montante,
                       perc_oneri, includi_oneri_in_iva, perc_iva, NULL, 0, 0, 0
                       FROM voci WHERE qe_id=? ORDER BY codice_completo"""
            cur.execute(query, (qid,))
            
        rows = cur.fetchall()
        montante = 0.0
        for r in rows:
            try: f_mont = r[5] 
            except: f_mont = 0
            if r[4] == 0 and f_mont == 1: montante += r[3]

        for r in rows:
            imp = r[3] if r[4] == 0 else (montante * r[3] / 100)
            oneri = imp * r[6] / 100
            base_iva = (imp + oneri) if r[7] == 1 else imp
            iva = base_iva * r[8] / 100
            tot_lordo = imp + oneri + iva

            forn = r[9] if r[9] else "Da individuare"
            try: a1 = float(r[10]) if r[10] else 0.0
            except: a1 = 0.0
            try: a2 = float(r[11]) if r[11] else 0.0
            except: a2 = 0.0
            
            if versione_id is None: a3 = tot_lordo 
            else: 
                try: a3 = float(r[12]) if r[12] is not None else (tot_lordo - a1 - a2)
                except: a3 = 0.0

            diff = abs(tot_lordo - (a1 + a2 + a3))
            icon = "✔" if diff < 0.02 else "⚠"
            tag = "ok" if diff < 0.02 else "err"

            self.tr.insert("", "end", iid=str(r[0]), values=(r[1], r[2], forn, self.fmt(tot_lordo), icon, a1, a2, a3), tags=(tag,))

    def on_select_voce(self, e):
        sel = self.tr.selection()
        if not sel: return
        self.voce_corrente_iid = sel[0]
        self.voce_corrente_id = int(sel[0])
        vals = self.tr.item(sel)['values']
        
        self.lbl_info_voce.config(text=f"{vals[0]} - {vals[1]}")
        self.totale_voce_target = self.parse(vals[3])
        self.var_forn.set(vals[2])
        
        self.vars_anni[0].set(self.fmt(vals[5]))
        self.vars_anni[1].set(self.fmt(vals[6]))
        self.vars_anni[2].set(self.fmt(vals[7]))
        self.calcola_dinamica()

    def calcola_dinamica(self, event=None):
        if not self.voce_corrente_id: return
        try:
            v1 = self.parse(self.vars_anni[0].get())
            v2 = self.parse(self.vars_anni[1].get())
            residuo = self.totale_voce_target - (v1 + v2)
            self.vars_anni[2].set(self.fmt(residuo))
            if residuo < -0.01: self.lbl_diff.config(text="Eccesso (A1+A2 > Tot)", foreground="red")
            else: self.lbl_diff.config(text="Bilanciato (Residuo in A3)", foreground="green")
        except: pass

    # =========================================================================
    # NUOVO METODO AGGIUNTO
    # =========================================================================
    def sposta_residuo_su_a1(self):
        if not self.voce_corrente_id: return
        
        # 1. Ottengo i valori attuali
        val_a1 = self.parse(self.vars_anni[0].get())
        val_a3 = self.parse(self.vars_anni[2].get()) # Questo è il residuo visibile
        
        # 2. Sommo il residuo (A3) al primo anno (A1)
        # Logica: Totale = A1 + A2 + A3. 
        # Se Nuovo_A1 = A1 + A3, allora il Nuovo_A3 calcolato dalla dinamica diventerà 0.
        nuovo_a1 = val_a1 + val_a3
        
        # 3. Aggiorno la GUI per A1
        self.vars_anni[0].set(self.fmt(nuovo_a1))
        
        # 4. Forzo il ricalcolo per aggiornare A3 (che diventerà 0) e le label di stato
        self.calcola_dinamica()
    # =========================================================================

    def applica_modifiche_riga(self):
        if not self.voce_corrente_iid: return
        f = self.var_forn.get()
        a1 = self.parse(self.vars_anni[0].get())
        a2 = self.parse(self.vars_anni[1].get())
        a3 = self.parse(self.vars_anni[2].get())
        diff = abs(self.totale_voce_target - (a1 + a2 + a3))
        icon = "✔" if diff < 0.02 else "⚠"
        tag = "ok" if diff < 0.02 else "err"
        curr = list(self.tr.item(self.voce_corrente_iid, 'values'))
        curr[2] = f; curr[4] = icon; curr[5] = a1; curr[6] = a2; curr[7] = a3
        self.tr.item(self.voce_corrente_iid, values=curr, tags=(tag,))
        self.tr.selection_set(self.voce_corrente_iid)

    def pulisci_form(self):
        self.voce_corrente_id = None
        self.lbl_info_voce.config(text="Seleziona una voce...")
        self.var_forn.set("")
        for v in self.vars_anni: v.set("0,00")
        self.lbl_diff.config(text="...")

    def salva_nuova_versione(self):
        qe_idx = self.cb_qe.current()
        if qe_idx == -1: return
        qid = self.map_qe[qe_idx]
        desc = simpledialog.askstring("Nuova Versione PF", "Inserisci una descrizione per questa versione:")
        if not desc: return
        ts = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        try:
            cur = self.conn.cursor()
            cur.execute("INSERT INTO fpv_testata (qe_id, descrizione, data_creazione) VALUES (?, ?, ?)", (qid, desc, ts))
            ver_id = cur.lastrowid
            items = self.tr.get_children()
            count = 0
            for iid in items:
                vals = self.tr.item(iid)['values']
                cur.execute("""INSERT INTO fpv_dettaglio (versione_id, voce_id, fornitore, anno_1, anno_2, anno_3)
                               VALUES (?, ?, ?, ?, ?, ?)""", (ver_id, int(iid), vals[2], vals[5], vals[6], vals[7]))
                count += 1
            self.conn.commit()
            messagebox.showinfo("Salvato", f"Versione salvata con successo!\nID: {ver_id}\nRighe: {count}")
            self.refresh_versioni()
        except Exception as e:
            self.conn.rollback(); messagebox.showerror("Errore", str(e))

    def esporta_excel(self):
        children = self.tr.get_children()
        if not children: return
        fn = filedialog.asksaveasfilename(defaultextension=".xlsx", title="Export FPV")
        if not fn: return
        wb = Workbook(); ws = wb.active; ws.title = "Piano Finanziario"
        ver_txt = self.cb_ver.get()
        ws['A1'] = f"PIANO FINANZIARIO - {ver_txt}"; ws['A1'].font = Font(bold=True, size=14); ws.append([])
        ws.append(["Codice", "Descrizione", "Fornitore", "Totale Lordo", "Anno 1", "Anno 2", "Anno 3"])
        for iid in children:
            v = self.tr.item(iid)['values']
            tot = self.parse(v[3])
            a1 = float(v[5])
            a2 = float(v[6])
            a3 = float(v[7])
            ws.append([v[0], v[1], v[2], tot, a1, a2, a3])
            for c in range(4, 8): ws.cell(row=ws.max_row, column=c).number_format = '#,##0.00 €'
        ws.column_dimensions['B'].width = 50; ws.column_dimensions['C'].width = 30
        wb.save(fn); messagebox.showinfo("OK", "File Excel creato."); self.app_root.apri_file(fn)
        
# =============================================================================
# APP MAIN
# =============================================================================
class CatalogoExportApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("QE Zero - Toolkit 3.x") 
        self.geometry("1000x700")
        self.db_path = self.trova_percorso_db()

        if not os.path.exists(self.db_path):
            messagebox.showwarning("Database non trovato", "Seleziona manualmente il file 'qe_zero.db'.")
            path_manuale = filedialog.askopenfilename(filetypes=[("SQLite DB", "*.db")])
            if path_manuale and os.path.exists(path_manuale): self.db_path = path_manuale
            else: self.destroy(); return

        self.setup_ui()
        self.avvia_connessioni()

    def trova_percorso_db(self):
        db_name = "qe_zero.db"
        program_dir = os.path.dirname(os.path.abspath(__file__))
        local_path = os.path.join(program_dir, "QE_DATI", db_name)
        if os.path.exists(local_path): return local_path
        base_docs = os.path.expanduser("~/Documents")
        return os.path.join(base_docs, "QE_DATI", db_name)

    def setup_ui(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tab1 = TabExportCataloghi(self.nb, self.db_path, self)
        self.nb.add(self.tab1, text="📂 Export Schede Catalogo")
        self.tab2 = TabBaseAsta(self.nb, self.db_path, self)
        self.nb.add(self.tab2, text="⚖️ Riepilogo Base d'Asta")
        self.tab3 = TabRiepilogoIva(self.nb, self.db_path, self)
        self.nb.add(self.tab3, text="💰 Riepilogo IVA")
        self.tab4 = TabCronoprogramma(self.nb, self.db_path, self)
        self.nb.add(self.tab4, text="📅 Gestione FPV")

    def avvia_connessioni(self):
        self.tab1.connetti_e_carica()
        self.tab2.connetti_e_carica()
        self.tab3.connetti_e_carica()
        self.tab4.connetti_e_carica()

    def apri_file(self, path):
        if os.name == 'nt': os.startfile(path)
        elif os.name == 'posix': subprocess.call(['open', path])

if __name__ == "__main__":
    app = CatalogoExportApp()
    app.mainloop()